import openpyxl
# from tabulate import tabulate

class DataFromExcel:

    def getTestData(self):

        Data = openpyxl.load_workbook("TestData\\PyExcelDemo.xlsx")

        activesheet = Data.active

        # cellData = activesheet.cell(row=1, column=2) # Read the cell data

        # activesheet.cell(row=2, column=2).value = "Hari" # Write the cell data

        list_of_dict = []

        for i in range(2, activesheet.max_row+1):

            if activesheet.cell(row=i, column=1).value == f"TestCase{i-1}":
                Dict = {}
                for j in range(2, activesheet.max_column+1):
                    Dict[activesheet.cell(row=1, column=j).value] = activesheet.cell(row=i, column=j).value
                       
                list_of_dict.append(Dict)
                
        print(list_of_dict)
        return list_of_dict