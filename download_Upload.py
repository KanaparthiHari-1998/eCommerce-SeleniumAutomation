from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl

driver  = webdriver.Chrome()
wait = WebDriverWait(driver, 10)
action = ActionChains(driver)
fruit_name = "//div[text()='{fruit_name}']"
file_path = r"C:\Users\Admin\Downloads\download.xlsx"
updated_Banana_Value = 390
Dict = {}

driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
wait.until(EC.visibility_of_element_located((By.ID, "downloadButton")))
driver.maximize_window()

driver.find_element(By.ID, "downloadButton").click()
time.sleep(3)

DataBook = openpyxl.load_workbook(file_path)
active_sheet = DataBook.active

for i in range(1, active_sheet.max_column+1):
    if active_sheet.cell(row=1, column=i).value == "price":
        Dict["col"] = i
        break

for i in range(1, active_sheet.max_row+1):
    for j in range(1, Dict["col"]+1):
        if active_sheet.cell(row=i, column=j).value == fruit_name:
            Dict["row"] = i
            break

active_sheet.cell(row=Dict["row"], column=Dict["col"]).value = updated_Banana_Value
DataBook.save(file_path)

driver.find_element(By.ID, "fileinput").send_keys(file_path)
wait.until(EC.invisibility_of_element_located((By.XPATH, "//div[text()='Updated Excel Data Successfully.']")))

price_col = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, f"//div[text()='{fruit_name}']/parent::div/parent::div/div[@id='cell-{price_col}-undefined']").text
print(price_col)
print(actual_price)
assert actual_price == str(updated_Banana_Value)

time.sleep(2)
driver.quit()